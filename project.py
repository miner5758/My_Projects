import time
import numpy as np
import openpyxl
import datetime
import pandas as pd
import matplotlib.pyplot as plt

t = 0
run_times = 1000
date_time_str = '05/07/21 01:55:19'
date_time_obj = datetime.datetime.strptime(date_time_str, '%d/%m/%y %H:%M:%S')


myRow = 1
myCol = 1
MyRow1 = 1
MyCol1 = 2
myRow2 = 1
myCol2 = 3
wbkName = 'game.xlsx'
wbk = openpyxl.load_workbook(wbkName)
sheet = wbk.get_sheet_by_name('Sheet1')
all_walks = []
go = False
while sheet.cell(row = myRow, column=myCol).value != None:
    myRow = myRow + 1
while sheet.cell(row = MyRow1, column=MyCol1).value != None:
    MyRow1 = MyRow1 + 1
while sheet.cell(row = myRow2, column=myCol2).value != None:
    myRow2 = myRow2 + 1

def countdown(t):
    go = True
    while t < 3600:
        print(t)
        t+=1
        time.sleep(1)
        global myRow
        for wks in wbk.worksheets:
            wks.cell(row=myRow, column=myCol).value = datetime.datetime.now()
        myRow = myRow + 1
        if go == True:
            game('t')
    if t == 3600:
        go = False
        for wks in wbk.worksheets:
            wks.cell(row=myRow, column=myCol).value = None
        wbk.save(wbkName)
        wbk.close
        plot('go')
    
    
def game(answer):
    random_walk = [0]
    for x in range (1000):
        step = random_walk[-1]
        dice = np.random.randint(1,7)
        if dice <= 2:
            step = max(0, step - 1)
        elif dice <= 5:
            step = step + 1
        else:
            step = step + np.random.randint(1,7)
        if np.random.rand() <= 0.001:
            step = 0
        random_walk.append(step)
    all_walks.append(random_walk)
    np_aw_t = np.transpose(np.array(all_walks))
    global MyRow1
    global myRow2
    global run_times
    MyRow1 = MyRow1 + 1
    for wks in wbk.worksheets:
        wks.cell(row=MyRow1-1, column=MyCol1).value = str(np_aw_t[-1,-1])
    for wks in wbk.worksheets:
        wks.cell(row=myRow2, column=myCol2).value = run_times    
    myRow2 = myRow2 + 1
    run_times = run_times + 1000
    
    
    
def plot(go_time):
    x = datetime.datetime.now().date()
    data = pd.read_excel(r"C:\Users\paula\game.xlsx")
    data = data[data['Timestamp'].dt.date == x]
    twohundred_and_below = data[data['Steps Reached'] <= 200]
    fourhundred_to_sixhundred = data[np.logical_and(data['Steps Reached']>=400, data['Steps Reached']<=600)]
    above = data[data['Steps Reached'] >= 600]
    new_data = {
        '200 and below': len(twohundred_and_below),
        '400 to 600': len(fourhundred_to_sixhundred),
        '600 and above': len(above)
        }
    d = pd.DataFrame(new_data,index=['Scores'])
    fig,axs = plt.subplots(figsize=(10,6))
    plt.title('Scores from random game test')
    plt.ylabel('Amount of times score amount was abotained')
    d.plot.bar(ax=axs)
    print(d)
countdown(int(t))