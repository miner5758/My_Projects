import sys
import random
import openpyxl
import pandas as pd
import os
import warnings
import datetime
import subprocess
import time as t
from dateutil import parser
import psutil
import string
import requests


class Zenex_manager():
    warnings.filterwarnings("ignore", category=DeprecationWarning) 
    if os.path.exists(os.path.join(os.getcwd(),'Zenex worker data.xlsx')):
            pass
    else:
        print('Downloading necessary files...')
        dls = 'https://github.com/miner5758/Miners_Projects/blob/Projects/Zenex_employee/Zenex%20worker%20data.xlsx?raw=true'
        resp = requests.get(dls)
        output = open('Zenex worker data.xlsx', 'wb')
        output.write(resp.content)
        output.close()
        
        dls = 'https://github.com/miner5758/Miners_Projects/raw/Projects/Zenex_employee/ZEO.txt'
        resp = requests.get(dls)
        output = open('ZEO.txt', 'wb')
        output.write(resp.content)
        output.close()
        
        dls = 'https://github.com/miner5758/Miners_Projects/raw/Projects/Zenex_employee/ansy.txt'
        resp = requests.get(dls)
        output = open('ansy.txt', 'wb')
        output.write(resp.content)
        output.close()
        
        dls = 'https://github.com/miner5758/Miners_Projects/raw/Projects/Zenex_employee/sec.txt'
        resp = requests.get(dls)
        output = open('sec.txt', 'wb')
        output.write(resp.content)
        output.close()
        
        dls = 'https://github.com/miner5758/Miners_Projects/raw/Projects/Zenex_employee/zenames.txt'
        resp = requests.get(dls)
        output = open('zenames.txt', 'wb')
        output.write(resp.content)
        output.close()
        
        dls = 'https://github.com/miner5758/Miners_Projects/raw/Projects/Zenex_employee/zenapi.txt'
        resp = requests.get(dls)
        output = open('zenapi.txt', 'wb')
        output.write(resp.content)
        output.close()
        
        dls = 'https://github.com/miner5758/Miners_Projects/raw/Projects/Zenex_employee/zenex_transfer.txt'
        resp = requests.get(dls)
        output = open('zenex_transfer.txt', 'wb')
        output.write(resp.content)
        output.close()
    with open('sec.txt') as f:
         sectors = f.read()
    sectors = sectors.split(',')
     
    wbk = openpyxl.load_workbook('Zenex worker data.xlsx')
    sheetvalue = ''
    sheet = ''
    
    
    def __init__(self,sector,api_key):
        if sector == '':
            raise Exception('you need to put which job sector your in! failed to load sector information')
        if sector.lower() in Zenex_manager.sectors:
            self.sector = sector.lower()
            Zenex_manager.sheetvalue = ('Zenex {0}_sector'.format(self.sector))
            Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
        else:
            raise Exception('{0} is not a job sector here! failed to load sector information'.format(sector))

        with open('zenapi.txt') as f:
            r = f.read()
        r = r.split('\n')
        for i in range(len(r)):
            for q in r:
                q = q.split(' ')
                e = q[0]
                q = q[1]
                if sector.lower() == e.strip(':').lower():
                    if api_key != str(q):
                        raise Exception('Invalid code, access denied')
                else:
                    pass

            
    def create_employee(self,name,position,race,salary,gender,date_created=datetime.datetime.now().date()):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
#-------------------Checks if excel is open and whether it should close it or not-------------------------       
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
#-------------------Checks if excel is open and whether it should close it or not--------------------------       

#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        if ',' in name or '.' in name:
            raise Exception('No periods or commas allowed in naming!')
        elif ',' in position or '.' in position:
            raise Exception('No periods or commas allowed in naming!')
        elif ',' in race or '.' in race:
            raise Exception('No periods or commas allowed in naming!')
        elif ',' in gender or '.' in gender:
            raise Exception('No periods or commas allowed in naming!')
        r = type(self)
        r = str(r)
        r = r = r.split('.')
        r = r[-1].replace("'", "")
        r = r.replace('>','')
        c = False
        yey = ''
        while not c:
            if r == 'Zenex_CEO':
                if date_created == 'NewSEctor':
                    print('yay')
                    yey = 'yes'
                    self.is_manager = 'Executive'
                    date_created = datetime.datetime.now().date()
                    c = True
                p = input('Would you like to create this employee as a executive?(Yes or No)')
                if p.lower() == 'yes':
                    Zenex_manager.sheet.insert_rows(idx=2,amount=1)
                    yey = 'yes'
                    self.is_manager = 'Executive'
                    c = True
                elif p.lower() == 'no':
                    c = True
                else:
                    print('Answer must be Yes or No')
            else:
                c = True
        if yey == 'yes':
            file = pd.ExcelFile('Zenex worker data.xlsx')
            data = pd.read_excel(file, Zenex_manager.sheetvalue)
            data = data[data["Is Manager"] == 'Executive']
            if name in data['Name of employee'].values:
                raise Exception('there is already an Executive with the name {0}, try adding a last name too'.format(name))
            if int(salary) > 100 or int(salary) < 60:
                raise Exception("{0} isn't a valid salary for a company Executive, failed to create profile".format(str(salary)))
            else:
                self.salary = int(salary)
        else:
            if int(salary) > 40 or int(salary) < 14:
                raise Exception("{0} isn't a valid salary for this company, failed to create profile".format(str(salary)))
            else:
                self.salary = int(salary)
        try:
            if (type(date_created) ==  datetime.date) == False:
                date_created = parser.parse(date_created)
                date_created = date_created.date()
        except TypeError:
            raise Exception('This is not a valid date')
        ID_row = 2
        ID_column = 1
        name_row = 2
        name_column = 2
        salary_row = 2
        salary_column = 3
        position_row = 2
        position_column = 4
        race_row = 2
        race_column = 5
        date_row = 2
        date_column = 6
        worked_row = 2
        worked_column = 7
        increase_row = 2
        increase_column = 8
        gender_row = 2
        gender_column = 9
        manager_row = 2
        manager_column = 10
        while Zenex_manager.sheet.cell(row = ID_row, column=ID_column).value != None:
            ID_row = ID_row + 1
        rows = [[name_row,name_column],[salary_row,salary_column],[position_row,position_column],
            [race_row,race_column],[date_row,date_column],[worked_row,worked_column],[increase_row,increase_column],
            [gender_row,gender_column],[manager_row,manager_column]]
        for i in rows:
            while Zenex_manager.sheet.cell(row = i[0], column=i[1]).value != None:
                rows[rows.index(i)][0]+=1
        self.name = name
        self.date_created = date_created
        self.position = position
        self.race = race
        self.worked = 0
        self.increase = 0
        self.gender = gender
        i = False
        if yey != 'yes':
            if r == 'Zenex_Xmanager' or r == 'Zenex_CEO':
                while not i:
                    p = input('Would you like to create this employee as a manager?(Yes or No)')
                    if p.lower() == 'yes':
                        self.is_manager = 'Yes'
                        i = True
                    elif p.lower() == 'no':
                        self.is_manager = 'No'
                        i = True
                    else:
                        print('Answer must be Yes or No')
            else:
                self.is_manager = 'No'
        else:
            if r == 'Zenex_Xmanager':
                while not i:
                    p = input('Would you like to create this employee as a manager?(Yes or No)')
                    if p.lower() == 'yes':
                        self.is_manager = 'Yes'
                        i = True
                    elif p.lower() == 'no':
                        self.is_manager = 'No'
                        i = True
                    else:
                        print('Answer must be Yes or No')
        values = [self.name,self.salary,self.position,self.race,self.date_created,self.worked,self.increase,self.gender,self.is_manager]
        password = ''.join((random.choice('123456789') for i in range(8)))
        global wbk
        Zenex_manager.sheet.cell(row=ID_row, column=ID_column).value = int(password)
        ic = 0
        for r in rows:
            if ic != 9:
                Zenex_manager.sheet.cell(row=r[0], column=r[1]).value = values[ic]        
                ic += 1
        
        Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
        Zenex_manager.wbk.close()
        rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
        for row in reversed(rows):
            if row[0].row == 1:
                break
            if row[0].value == None:
                Zenex_manager.sheet.delete_rows(row[0].row, 1)
        Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
        Zenex_manager.wbk.close()
        if yey == 'yes':
            with open('zenames.txt') as f:
                r = f.read()
            r = r.split('\n')
            with open('zenames.txt','w') as f:
                for i in r:
                    if self.sector.capitalize() in i:
                        if i == r[0]:
                            f.write(i+'.'+self.name+',Void')
                        else:
                            f.write('\n'+i+'.'+self.name+',Void')
                    else:
                        if i == r[0]:
                            f.write(i)
                        else:
                            f.write('\n'+i)
                
                
        
        
    def select_by_ID(self,ID):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      

#------checks if there is an employee here by inputted ID, whether it should use the selected employee or not, and as well as if the employee is a manager------------------------
        file = pd.ExcelFile('Zenex worker data.xlsx')
        data = pd.read_excel(file, Zenex_manager.sheetvalue)
        if int(ID) not in data['employee ID'].values:
            raise Exception('There is no employee by that ID')
        data = data[data["employee ID"] == int(ID)]
        d = data['Is Manager'].values
        d = ' '.join(map(str,d))
        r = type(self)
        r = str(r)
        r = r = r.split('.')
        r = r[-1].replace("'", "")
        r = r.replace('>','')
        if d == 'Yes' or d == 'Executive': 
            if r == 'Zenex_Xmanager':
                if d == 'Yes':
                    pass
                elif d == 'Executive':
                    raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')                
            elif r == 'Zenex_CEO':
                pass
            else:
                raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')                      
#------checks if there is an employee here by inputted ID, whether it should use the selected employee or not, and as well as if the employee is a manager------------------------
                
        
        self.is_manager = d
        self.ID = ID
        self.name = data['Name of employee'].values
        self.name = ' '.join(map(str,self.name))
        self.race = data['Race/ethnicity'].values
        self.race = ' '.join(map(str,self.race))
        self.salary = data['$ Salary per/hour $'].values
        self.salary = ' '.join(map(str,self.salary))
        self.salary = int(self.salary)
        self.position = data['Position of employee'].values
        self.position = ' '.join(map(str,self.position))
        self.increase = data['Total Salary Increase'].values
        self.increase = ' '.join(map(str,self.increase))
        self.increase = int(self.increase)
        self.gender = data['Gender'].values
        self.gender = ' '.join(map(str,self.gender))
        self.worked = data['Hours Worked'].values
        self.worked = ' '.join(map(str,self.worked))
        self.worked = int(self.worked)
        self.date_created = data['Date employed'].values
        self.date_created = ' '.join(map(str,self.date_created))
        self.date_created = parser.parse(self.date_created)
        self.date_created = self.date_created.date()
        
        
    def select_by_name(self,name):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      

#------checks if there is an employee here by inputted name, whether it should use the selected employee or not, and as well as if the employee is a manager------------------------
        file = pd.ExcelFile('Zenex worker data.xlsx')
        data = pd.read_excel(file, Zenex_manager.sheetvalue)
        if name not in data['Name of employee'].values:
            raise Exception('There is no employee by the name {0} (this is case sensitive so make sure your typing it in right)'.format(name))
        data = data[data['Name of employee'] == str(name)]
        if len(data) != 1:
            raise Exception('there are 2 employee by the name {0}, try searching by employee ID'.format(name))
        d = data['Is Manager'].values
        d = ' '.join(map(str,d))
        r = type(self)
        r = str(r)
        r = r = r.split('.')
        r = r[-1].replace("'", "")
        r = r.replace('>','')
        if d == 'Yes' or d == 'Executive': 
            if r == 'Zenex_Xmanager':
                if d == 'Yes':
                    pass
                elif d == 'Executive':
                    raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')                
            elif r == 'Zenex_CEO':
                pass
            else:
                raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')      
#------checks if there is an employee here by inputted name, whether it should use the selected employee or not, and as well as if the employee is a manager------------------------

        self.is_manager = d
        self.ID = data['employee ID'].values
        self.ID = ' '.join(map(str,self.ID))
        self.name = name
        self.race = data['Race/ethnicity'].values
        self.race = ' '.join(map(str,self.race))
        self.salary = data['$ Salary per/hour $'].values
        self.salary = ' '.join(map(str,self.salary))
        self.salary = int(self.salary)
        self.position = data['Position of employee'].values
        self.position = ' '.join(map(str,self.position))
        self.increase = data['Total Salary Increase'].values
        self.increase = ' '.join(map(str,self.increase))
        self.increase = int(self.increase)
        self.gender = data['Gender'].values
        self.gender = ' '.join(map(str,self.gender))
        self.worked = data['Hours Worked'].values
        self.worked = ' '.join(map(str,self.worked))
        self.worked = int(self.worked)
        self.date_created = data['Date employed'].values
        self.date_created = ' '.join(map(str,self.date_created))
        self.date_created = parser.parse(self.date_created)
        self.date_created = self.date_created.date()

    


    def delete_employee_ID(self,ID=0):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
        yey = ''
        op = ''
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      

#-------------------Checks if excel is open and whether it should close it or not------------------------       
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
#-------------------Checks if excel is open and whether it should close it or not------------------------       

#------checks if there is an employee here by inputted ID, whether it should use the selected employee or not, and as well as if the employee is a manager------------------------
        try:
            if ID == 0:
                ID = self.ID
        except AttributeError:
            raise Exception('you need to specify which employee you wanna delete, either by doing emloyee.select_by_ID(ID) then employee.delete_employee_ID(ID) or employee.delete_employee_ID(ID)')
        file = pd.ExcelFile('Zenex worker data.xlsx')
        data = pd.read_excel(file, Zenex_manager.sheetvalue)
        if int(ID) not in data['employee ID'].values:
            raise Exception('There is no employee by that ID')
        data = data[data["employee ID"] == int(ID)]
        d = data['Is Manager'].values
        d = ' '.join(map(str,d))
        r = type(self)
        r = str(r)
        r = r = r.split('.')
        r = r[-1].replace("'", "")
        r = r.replace('>','')
        if d == 'Yes' or d == 'Executive': 
            if r == 'Zenex_Xmanager':
                if d == 'Yes':
                    pass
                elif d == 'Executive':
                    raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')                
            elif r == 'Zenex_CEO':
                if d == 'Executive':
                    yey = 'yes'
                    op = data['Name of employee'].values
                    op = ' '.join(map(str,op))
                else:
                    pass
            else:
                raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')      
#------checks if there is an employee here by inputted ID, whether it should use the selected employee or not, and as well as if the employee is a manager------------------------

        rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
        for row in reversed(rows):
            if row[0].row == 1:
                break
            if int(row[0].value) == int(ID):
                Zenex_manager.sheet.delete_rows(row[0].row, 1)
        Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
        Zenex_manager.wbk.close()
        if yey == 'yes':
            with open('zenames.txt') as f:
                yu = f.readlines()
            with open('zenames.txt','w') as f:
                for i in yu:
                    if self.sector.capitalize() in i:
                        i = i.strip(' '+self.sector.capitalize()+':')
                        i = i.split('.')
                        for w in i:
                            if op in w:
                                r = i.index(w)
                                del i[r]
                        fr = ''
                        for te in i:
                            if te != i[-1]:
                                fr = fr + te + '.'
                            else:
                                fr = fr + te
                        fr = self.sector.capitalize() + ': ' + fr
                        if '.' not in fr:
                            fr = fr + '\n'
                        f.write(fr)
                    else:
                        f.write(i.strip('\n')+'\n')
               
        
                    
        try:
            if ID == self.ID:
                del self.is_manager
                del self.ID
                del self.name
                del self.race
                del self.salary
                del self.position
                del self.increase
                del self.gender
                del self.worked
                del self.date_created
        except AttributeError:
            pass
        
        
        
    def delete_employee_name(self,name=''):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
        yey = ''
        op = '' 
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      

#-------------------Checks if excel is open and whether it should close it or not------------------------       
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
#-------------------Checks if excel is open and whether it should close it or not------------------------       

#------checks if there is an employee here by inputted name, whether it should use the selected employee or not, and as well as if the employee is a manager------------------------
        try:
            if name == '':
                name = self.name
        except AttributeError:
            raise Exception('you need to specify which employee you wanna delete, either by doing emloyee.select_by_name(name) then employee.delete_employee_name(name) or employee.delete_employee_name(name)')
            
        file = pd.ExcelFile('Zenex worker data.xlsx')
        data = pd.read_excel(file, Zenex_manager.sheetvalue)
        if name not in data['Name of employee'].values:
            raise Exception('There is no employee by the name {0} (this is case sensitive so make sure your typing it in right)'.format(name))
        data = data[data['Name of employee'] == str(name)]
        if len(data) != 1:
            raise Exception('there are 2 employee by the name {0}, try using employee ID'.format(name))
        d = data['Is Manager'].values
        d = ' '.join(map(str,d))
        r = type(self)
        r = str(r)
        r = r = r.split('.')
        r = r[-1].replace("'", "")
        r = r.replace('>','')
        if d == 'Yes' or d == 'Executive': 
            if r == 'Zenex_Xmanager':
                if d == 'Yes':
                    pass
                elif d == 'Executive':
                    raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')                
            elif r == 'Zenex_CEO':
                if d == 'Executive':
                    yey = 'yes'
                    op = data['Name of employee'].values
                    op = ' '.join(map(str,op))
                else:
                    pass
            else:
                raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')         
#------checks if there is an employee here by inputted name, whether it should use the selected employee or not, and as well as if the employee is a manager------------------------

        rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
        for row in reversed(rows):
            #print(row[0].value)
            if row[1].row == 1:
                break
            if str(row[1].value) == str(name):
                Zenex_manager.sheet.delete_rows(row[1].row, 1)
        Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
        Zenex_manager.wbk.close()
        if yey == 'yes':
            with open('zenames.txt') as f:
                yu = f.readlines()
            with open('zenames.txt','w') as f:
                for i in yu:
                    if self.sector.capitalize() in i:
                        i = i.strip(' '+self.sector.capitalize()+':')
                        i = i.split('.')
                        for w in i:
                            if op in w:
                                r = i.index(w)
                                del i[r]
                        fr = ''
                        for te in i:
                            if te != i[-1]:
                                fr = fr + te + '.'
                            else:
                                fr = fr + te
                        fr = self.sector.capitalize() + ': ' + fr
                        if '.' not in fr:
                            fr = fr + '\n'
                        f.write(fr)
                    else:
                        f.write(i.strip('\n')+'\n')
        
        try:
            if name == self.name:
                del self.is_manager
                del self.ID
                del self.name
                del self.race
                del self.salary
                del self.position
                del self.increase
                del self.gender
                del self.worked
                del self.date_created
        except AttributeError:
            pass        
    
    def selected_employee(self):
        try:
            print('Name:',self.name)
            print('Employee ID:', self.ID)
            print('Race:',self.race)
            print('Gender:',self.gender)
            print('Salary:',self.salary)
            print('Salary increase:',self.increase)
            print('Position:',self.position)
            print('Hours worked:',self.worked)
            print('Date employeed:',self.date_created)
            print('sector:',self.sector)
            print('Is manager:',self.is_manager)
        except AttributeError:
            raise Exception('No employee selected at the moment')
            
            
            
            
            
    def give_raise_name(self,amount,name=''):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      

#-------------------Checks if excel is open and whether it should close it or not------------------------       
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
#-------------------Checks if excel is open and whether it should close it or not------------------------       

#------checks if there is an employee here by inputted name, whether it should use the selected employee or not, whether there are 2 employees by 1 name,as well as if the employee is a manager------------------------  
        try:
            if name == '':
                name = self.name
                newcrease = self.salary + amount
                newsal = self.salary + amount
        except AttributeError:
            raise Exception('you need to specify which employee you wanna give a raise, either by doing emloyee.select_by_name(name) then employee.give_raise_name() or employee.give_raise_name(name)')
            
            
        file = pd.ExcelFile('Zenex worker data.xlsx')
        data = pd.read_excel(file, Zenex_manager.sheetvalue)
        if name not in data['Name of employee'].values:
            raise Exception('There is no employee by the name {0} (this is case sensitive so make sure your typing it in right)'.format(name))
        data = data[data['Name of employee'] == str(name)]
        if len(data) != 1:
            raise Exception('there are 2 employee by the name {0}, try searching by employee ID'.format(name))
        d = data['Is Manager'].values
        d = ' '.join(map(str,d))
        r = type(self)
        r = str(r)
        r = r = r.split('.')
        r = r[-1].replace("'", "")
        r = r.replace('>','')
        yey = ''
        if d == 'Yes' or d == 'Executive': 
            if r == 'Zenex_Xmanager':
                if d == 'Yes':
                    pass
                elif d == 'Executive':
                    raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')                
            elif r == 'Zenex_CEO':
                if d == 'Executive':
                    yey = 'yes'
            else:
                raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')                      
#------checks if there is an employee here by inputted name, whether it should use the selected employee or not, whether there are 2 employees by 1 name,as well as if the employee is a manager------------------------


        for row in Zenex_manager.sheet.iter_rows():
            for cell in row:
                if cell.value == name:
                    Zenex_manager.sheet.cell(row=cell.row, column=3).value += amount
                    Zenex_manager.sheet.cell(row=cell.row, column=8).value += amount
                    if yey == 'yes':
                        if Zenex_manager.sheet.cell(row=cell.row, column=3).value > 100:
                            raise Exception('its impossible for an Executive to have a salary higher than 100')
                    else:
                        if Zenex_manager.sheet.cell(row=cell.row, column=3).value > 40:
                            raise Exception('its impossible for an employee to have a salary higher than 40')
                    try:
                        self.salary = newsal
                        self.increase = newcrease
                    except UnboundLocalError:
                        pass
        Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
        Zenex_manager.wbk.close()       
        
        
        
    def give_paycut_name(self,amount,name=''):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      

#-------------------Checks if excel is open and whether it should close it or not------------------------       
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
#-------------------Checks if excel is open and whether it should close it or not------------------------       

#------checks if there is an employee here by inputted name, whether it should use the selected employee or not, whether there are 2 employees by 1 name,as well as if the employee is a manager------------------------
        try:
            if name == '':
                name = self.name
                newcrease = self.increase - amount
                newsal = self.salary - amount
        except AttributeError:
            raise Exception('you need to specify which employee you wanna give a paycut, either by doing emloyee.select_by_name(name) then employee.give_paycut_name() or employee.give_paycut_name(name)')
        file = pd.ExcelFile('Zenex worker data.xlsx')
        data = pd.read_excel(file, Zenex_manager.sheetvalue)
        if name not in data['Name of employee'].values:
            raise Exception('There is no employee by the name {0} (this is case sensitive so make sure your typing it in right)'.format(name))
        data = data[data['Name of employee'] == str(name)]
        if len(data) != 1:
            raise Exception('there are 2 employee by the name {0}, try searching by employee ID'.format(name))
        d = data['Is Manager'].values
        d = ' '.join(map(str,d))
        r = type(self)
        r = str(r)
        r = r = r.split('.')
        r = r[-1].replace("'", "")
        r = r.replace('>','')
        yey = ''
        if d == 'Yes' or d == 'Executive': 
            if r == 'Zenex_Xmanager':
                if d == 'Yes':
                    pass
                elif d == 'Executive':
                    raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')                
            elif r == 'Zenex_CEO':
                if d == 'Executive':
                    yey = 'yes'
            else:
                raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')                      
#------checks if there is an employee here by inputted name, whether it should use the selected employee or not, whether there are 2 employees by 1 name,as well as if the employee is a manager------------------------

        for row in Zenex_manager.sheet.iter_rows():
            for cell in row:
                if cell.value == name:
                    Zenex_manager.sheet.cell(row=cell.row, column=3).value -= amount
                    Zenex_manager.sheet.cell(row=cell.row, column=8).value -= amount
                    if yey == 'yes':
                        if Zenex_manager.sheet.cell(row=cell.row, column=3).value < 60:
                            raise Exception('its impossible for an Executive to have a salary lower than 60')
                    else:
                        if Zenex_manager.sheet.cell(row=cell.row, column=3).value < 14:
                            raise Exception('its impossible for an employee to have a salary lower than 14')
                    try:
                        self.salary = newsal
                        self.increase = newcrease
                    except UnboundLocalError:
                        pass
        Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
        Zenex_manager.wbk.close()
        
    def give_raise_ID(self,amount,ID=0):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      

#-------------------Checks if excel is open and whether it should close it or not------------------------       
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
#-------------------Checks if excel is open and whether it should close it or not------------------------       
        
#------checks if there is an employee here by inputted ID, whether it should use the selected employee or not, and as well as if the employee is a manager------------------------
        try:
            if ID == 0:
                ID = self.ID
                newcrease = self.increase + amount
                newsal = self.salary + amount
        except AttributeError:
            raise Exception('you need to specify which employee you wanna give a raise, either by doing emloyee.select_by_ID(ID) then employee.give_raise_ID() or employee.give_raise_ID(ID)')
        file = pd.ExcelFile('Zenex worker data.xlsx')
        data = pd.read_excel(file, Zenex_manager.sheetvalue)
        if int(ID) not in data['employee ID'].values:
            raise Exception('There is no employee by that ID')
        data = data[data["employee ID"] == int(ID)]
        d = data['Is Manager'].values
        d = ' '.join(map(str,d))
        r = type(self)
        r = str(r)
        r = r = r.split('.')
        r = r[-1].replace("'", "")
        r = r.replace('>','')
        yey = ''
        if d == 'Yes' or d == 'Executive': 
            if r == 'Zenex_Xmanager':
                if d == 'Yes':
                    pass
                elif d == 'Executive':
                    raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')                
            elif r == 'Zenex_CEO':
                if d == 'Executive':
                    yey = 'yes'
            else:
                raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')                      
#------checks if there is an employee here by inputted ID, whether it should use the selected employee or not, and as well as if the employee is a manager------------------------

        rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
        for row in reversed(rows):
            if row[0].row == 1:
                break
            if int(row[0].value) == int(ID):
                Zenex_manager.sheet.cell(row=row[0].row, column=3).value += amount
                Zenex_manager.sheet.cell(row=row[0].row, column=8).value += amount
                if yey == 'yes':
                    if Zenex_manager.sheet.cell(row=row[0].row, column=3).value > 100:
                       raise Exception('its impossible for an Executive to have a salary higher than 100')
                else:
                    if Zenex_manager.sheet.cell(row=row[0].row, column=3).value > 40:
                       raise Exception('its impossible for an employee to have a salary higher than 40')
                try:
                    self.salary = newsal
                    self.increase = newcrease
                except UnboundLocalError:
                    pass
        Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
        Zenex_manager.wbk.close()
       
        
    
    def give_paycut_ID(self,amount,ID=0):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      

#-------------------Checks if excel is open and whether it should close it or not------------------------       
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
#-------------------Checks if excel is open and whether it should close it or not------------------------       
        
#------checks if there is an employee here by inputted ID, whether it should use the selected employee or not, and as well as if the employee is a manager------------------------
        try:
            if ID == 0:
                ID = self.ID
                newcrease = self.increase = amount
                newsal = self.salary + amount
        except AttributeError:
            raise Exception('you need to specify which employee you wanna give a paycut, either by doing emloyee.select_by_ID(ID) then employee.give_paycut_ID() or employee.give_paycut_ID(ID)')
        file = pd.ExcelFile('Zenex worker data.xlsx')
        data = pd.read_excel(file, Zenex_manager.sheetvalue)
        if int(ID) not in data['employee ID'].values:
            raise Exception('There is no employee by that ID')
        data = data[data["employee ID"] == int(ID)]
        d = data['Is Manager'].values
        d = ' '.join(map(str,d))
        r = type(self)
        r = str(r)
        r = r = r.split('.')
        r = r[-1].replace("'", "")
        r = r.replace('>','')
        yey = ''
        if d == 'Yes' or d == 'Executive': 
            if r == 'Zenex_Xmanager':
                if d == 'Yes':
                    pass
                elif d == 'Executive':
                    raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')                
            elif r == 'Zenex_CEO':
                if d == 'Executive':
                    yey = 'yes'
            else:
                raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')                     
#------checks if there is an employee here by inputted ID, whether it should use the selected employee or not, and as well as if the employee is a manager------------------------

        rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
        for row in reversed(rows):
            if row[0].row == 1:
                break
            if int(row[0].value) == int(ID):
                Zenex_manager.sheet.cell(row=row[0].row, column=3).value -= amount
                Zenex_manager.sheet.cell(row=row[0].row, column=8).value -= amount
                if yey == 'yes':
                    if Zenex_manager.sheet.cell(row=row[0].row, column=3).value < 60:
                        raise Exception('its impossible for an Executive to have a salary lower than 60')
                else:
                    if Zenex_manager.sheet.cell(row=row[0].row, column=3).value < 14:
                        raise Exception('its impossible for an employee to have a salary lower than 14')
                try:
                    self.salary = newsal
                    self.increase = newcrease
                except UnboundLocalError:
                    pass
        Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
        Zenex_manager.wbk.close()
        
        

        
    def promote_name(self,name=''):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
        val = 'Yes'
        np = ''
        gp = ''
        eman = ''
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      

#-------------------Checks if excel is open and whether it should close it or not------------------------       
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
#-------------------Checks if excel is open and whether it should close it or not------------------------       

#------checks if there is an employee here by inputted name, whether it should use the selected employee or not, whether there are 2 employees by 1 name,as well as if the employee is a manager------------------------
        try:
            if name == '':
                name = self.name
        except AttributeError:
            raise Exception('you need to specify which employee you wanna promote, either by doing emloyee.select_by_name(name) then employee.promote_name() or employee.promote_name(name)')
        file = pd.ExcelFile('Zenex worker data.xlsx')
        data = pd.read_excel(file, Zenex_manager.sheetvalue)
        if name not in data['Name of employee'].values:
            raise Exception('There is no employee by the name {0} (this is case sensitive so make sure your typing it in right)'.format(name))
        data = data[data['Name of employee'] == str(name)]
        if len(data) != 1:
            raise Exception('there are 2 employee by the name {0}, try searching by employee ID'.format(name))
        d = data['Is Manager'].values
        d = ' '.join(map(str,d))
        r = type(self)
        r = str(r)
        r = r = r.split('.')
        r = r[-1].replace("'", "")
        r = r.replace('>','')
        if d == 'Yes' or d == 'Executive': 
            if r == 'Zenex_Xmanager':
                if d == 'Yes':
                    pass
                elif d == 'Executive':
                    raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')                
            elif r == 'Zenex_CEO':
                if d == 'Executive':
                    raise Exception('This person is a Executive, you dont really get better than that')
                else:
                    g = False
                    while not g:
                        ex = input('Want to make this guy an Executive?')
                        if ex.lower() == 'yes':
                            val = 'Executive'
                            np = 60
                            gp = (np - int(data['$ Salary per/hour $'].values)) + int(data['Total Salary Increase'].values)
                            g = True
                        elif ex.lower() == 'no':
                            g = True
                        else:
                            print('Not a valid answer, either yes or no')
            else:
                raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')             
#------checks if there is an employee here by inputted name, whether it should use the selected employee or not, whether there are 2 employees by 1 name,as well as if the employee is a manager------------------------

        for row in Zenex_manager.sheet.iter_rows():
            for cell in row:
                if cell.value == name:
                    Zenex_manager.sheet.cell(row=cell.row, column=10).value = val
                    if np != '':
                        atad = pd.read_excel(file, Zenex_manager.sheetvalue)
                        atad = atad[atad['Name of employee'] == str(name)]
                        atad = atad[atad['Is Manager'] == 'Executive']
                        if len(atad) >= 1:
                            lk = False
                            while not lk:
                                eman = input('Since there already is an Exectuive with the Name {0} try adding a last name or something')
                                if eman == name:
                                    print('Cant be the same as the old one')
                                else:
                                    lk = True
                        lio = []
                        ew = 1
                        sd = 0
                        Zenex_manager.sheet.cell(row=cell.row, column=3).value = int(np)
                        Zenex_manager.sheet.cell(row=cell.row, column=8).value = int(gp)
                        Zenex_manager.sheet.insert_rows(2)
                        for i in range(1,11):
                            lio.append(Zenex_manager.sheet.cell(row=cell.row, column=i).value)
                        if eman != '':
                            lio[1] == eman
                        rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
                        for row in reversed(rows):
                            #print(row[0].value)
                            if row[1].row == 1:
                                break
                            if str(row[1].value) == str(name):
                                Zenex_manager.sheet.delete_rows(row[1].row, 1)
                        for row in Zenex_manager.sheet.iter_rows():
                            for cell in row:
                                if cell.value == None:
                                    Zenex_manager.sheet.cell(row=cell.row, column=ew).value = lio[sd]
                                    ew+=1
                                    sd+=1
                            
        rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
        for row in reversed(rows):
            if row[0].row == 1:
                break
            if row[0].value == None:
                Zenex_manager.sheet.delete_rows(row[0].row, 1)
        try:
            if name == self.name:
                del self.is_manager
                del self.ID
                del self.name
                del self.race
                del self.salary
                del self.position
                del self.increase
                del self.gender
                del self.worked
                del self.date_created
        except AttributeError:
            pass
        Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
        Zenex_manager.wbk.close()
        if val == 'Executive':
            with open('zenames.txt') as f:
                r = f.read()
            r = r.split('\n')
            with open('zenames.txt','w') as f:
                if eman != '':
                    name = eman
                for i in r:
                    if self.sector.capitalize() in i:
                        if i == r[0]:
                            f.write(i+'.'+name+',Void')
                        else:
                            f.write('\n'+i+'.'+name+',Void')
                    else:
                        if i == r[0]:
                            f.write(i)
                        else:
                            f.write('\n'+i)
        
        
    def promote_ID(self,ID=0):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
        val = 'Yes'
        np = ''
        gp = ''
        eman = ''
        op = ''
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      

#-------------------Checks if excel is open and whether it should close it or not------------------------       
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
#-------------------Checks if excel is open and whether it should close it or not------------------------       

#------checks if there is an employee here by inputted ID, whether it should use the selected employee or not, and as well as if the employee is a manager------------------------
        try:
            if ID == 0:
                ID = self.ID
        except AttributeError:
            raise Exception('you need to specify which employee you wanna promote, either by doing emloyee.select_by_ID(ID) then employee.promote_ID() or employee.promote_ID(ID)')
        file = pd.ExcelFile('Zenex worker data.xlsx')
        data = pd.read_excel(file, Zenex_manager.sheetvalue)
        if int(ID) not in data['employee ID'].values:
            raise Exception('There is no employee by that ID')
        data = data[data["employee ID"] == int(ID)]
        d = data['Is Manager'].values
        d = ' '.join(map(str,d))
        r = type(self)
        r = str(r)
        r = r = r.split('.')
        r = r[-1].replace("'", "")
        r = r.replace('>','')
        if d == 'Yes' or d == 'Executive': 
            if r == 'Zenex_Xmanager':
                if d == 'Yes':
                    pass
                elif d == 'Executive':
                    raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')                
            elif r == 'Zenex_CEO':
                if d == 'Executive':
                    raise Exception('This person is a Executive, you dont really get better than that')
                else:
                    g = False
                    while not g:
                        ex = input('Want to make this guy an Executive?')
                        if ex.lower() == 'yes':
                            val = 'Executive'
                            np = 60
                            gp = (np - int(data['$ Salary per/hour $'].values)) + int(data['Total Salary Increase'].values)
                            g = True
                        elif ex.lower() == 'no':
                            g = True
                        else:
                            print('Not a valid answer, either yes or no')
            else:
                raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')    
#------checks if there is an employee here by inputted name, whether it should use the selected employee or not, whether there are 2 employees by 1 name,as well as if the employee is a manager------------------------

        for row in Zenex_manager.sheet.iter_rows():
            for cell in row:
                if cell.value == ID:
                    Zenex_manager.sheet.cell(row=cell.row, column=10).value = val
                    if np != '':
                        op = data['Name of employee'].values
                        op = ' '.join(map(str,op))
                        atad = pd.read_excel(file, Zenex_manager.sheetvalue)
                        atad = atad[atad['Name of employee'] == str(op)]
                        atad = atad[atad['Is Manager'] == 'Executive']
                        if len(atad) >= 1:
                            lk = False
                            while not lk:
                                eman = input('Since there already is an Exectuive with the Name {0} try adding a last name or something')
                                if eman == op:
                                    print('Cant be the same as the old one')
                                else:
                                    lk = True
                            eman = str(eman)
                        lio = []
                        ew = 1
                        sd = 0
                        Zenex_manager.sheet.cell(row=cell.row, column=3).value = int(np)
                        Zenex_manager.sheet.cell(row=cell.row, column=8).value = int(gp)
                        Zenex_manager.sheet.insert_rows(2)
                        for i in range(1,11):
                            lio.append(Zenex_manager.sheet.cell(row=cell.row, column=i).value)
                        if eman != '':
                            lio[1] = eman
                        rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
                        for row in reversed(rows):
                            if row[0].row == 1:
                                break
                            try:
                                if int(row[0].value) == int(ID):
                                    Zenex_manager.sheet.delete_rows(row[1].row, 1)
                            except TypeError:
                                pass
                                
                        for row in Zenex_manager.sheet.iter_rows():
                            for cell in row:
                                if cell.value == None:
                                    Zenex_manager.sheet.cell(row=cell.row, column=ew).value = lio[sd]
                                    ew+=1
                                    sd+=1
                            
        rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
        for row in reversed(rows):
            if row[0].row == 1:
                break
            if row[0].value == None:
                Zenex_manager.sheet.delete_rows(row[0].row, 1)
        try:
            if ID == self.ID:
                del self.is_manager
                del self.ID
                del self.name
                del self.race
                del self.salary
                del self.position
                del self.increase
                del self.gender
                del self.worked
                del self.date_created
        except AttributeError:
            pass
        Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))    
        Zenex_manager.wbk.close()
        if val == 'Executive':
            with open('zenames.txt') as f:
                r = f.read()
            r = r.split('\n')
            with open('zenames.txt','w') as f:
                if eman != '':
                    op = eman
                for i in r:
                    if self.sector.capitalize() in i:
                        if i == r[0]:
                            f.write(i+'.'+str(op)+',Void')
                        else:
                            f.write('\n'+i+'.'+str(op)+',Void')
                    else:
                        if i == r[0]:
                            f.write(i)
                        else:
                            f.write('\n'+i)
        
        
        
        
        
        
    def update_information_name(self,name=''):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      

#-------------------Checks if excel is open and whether it should close it or not------------------------       
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
#-------------------Checks if excel is open and whether it should close it or not------------------------       

#------checks if there is an employee here by inputted name, whether it should use the selected employee or not, and as well as if the employee is a manager------------------------       
        try:
            if name == '':
                name = self.name
        except AttributeError:
            raise Exception('you need to specify which employee you wanna update, either by doing emloyee.select_by_name(name) then employee.update_information_name() or employee.update_information_name(name)')
        file = pd.ExcelFile('Zenex worker data.xlsx')
        data = pd.read_excel(file, Zenex_manager.sheetvalue)
        if name not in data['Name of employee'].values:
            raise Exception('There is no employee by the name {0} (this is case sensitive so make sure your typing it in right)'.format(name))
        data = data[data['Name of employee'] == str(name)]
        if len(data) != 1:
            raise Exception('there are 2 employee by the name {0}, try searching by employee ID'.format(name))
        d = data['Is Manager'].values
        d = ' '.join(map(str,d))
        r = type(self)
        r = str(r)
        r = r = r.split('.')
        r = r[-1].replace("'", "")
        r = r.replace('>','')
        if d == 'Yes' or d == 'Executive': 
            if r == 'Zenex_Xmanager':
                if d == 'Yes':
                    pass
                elif d == 'Executive':
                    raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')                
            elif r == 'Zenex_CEO':
                pass
            else:
                raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')         
#------checks if there is an employee here by inputted name, whether it should use the se;ected employee or not, and as well as if the employee is a manager------------------------
        
#------------------------ask you what you want to edit and then edits based on your input-----------------------------------------------------------
        options = {}
        options["Position"] = 1
        options["Race/ethnicity"] = 2
        options["Date employyed"] = 3
        options["Hours worked"] = 4
        options["Gender"] = 5
        options["Name"] = 6
        index = 0
        indexValidList = []
        print('which one do you want to edit?')
        for optionName in options:
            index = index + 1
            indexValidList.extend([options[optionName]])
            print(str(index) + ') ' + optionName)
        inputValid = False
        while not inputValid:
            inputRaw = input("option('must be a number'): ")
            inputNo = int(inputRaw) - 1
            if inputNo > -1 and inputNo < len(indexValidList):
                selected = indexValidList[inputNo]
                if selected == 1:
                    o = input('What would you like to replace this persons position with?')
                    for row in Zenex_manager.sheet.iter_rows():
                        for cell in row:
                            if cell.value == name:
                                self.position = o
                                Zenex_manager.sheet.cell(row=cell.row, column=4).value = str(o)
                                Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
                                Zenex_manager.wbk.close()
                elif selected == 2:
                    o = input('What would you like to replace this persons race with?')
                    for row in Zenex_manager.sheet.iter_rows():
                        for cell in row:
                            if cell.value == name:
                                self.race = o
                                Zenex_manager.sheet.cell(row=cell.row, column=5).value = str(o)
                                Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
                                Zenex_manager.wbk.close()
                elif selected == 3:
                   try:
                       o = input('What would you like to replace this persons date employeed with?')
                       if (type(o) ==  datetime.date) == False:
                           o = parser.parse(o)
                           o = o.date()
                   except TypeError:
                       raise Exception('This is not a valid date')
                   for row in Zenex_manager.sheet.iter_rows():
                       for cell in row:
                           if cell.value == name:
                               self.date_created = o
                               Zenex_manager.sheet.cell(row=cell.row, column=6).value = o  
                               Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
                               Zenex_manager.wbk.close()
                elif selected == 4:
                    try:
                        o = input('What would you like to replace this persons hours worked with?')
                        o = int(o)
                    except ValueError:
                        raise Exception('{0} is not a valid number for this'.format(o))
                    for row in Zenex_manager.sheet.iter_rows():
                           for cell in row:
                               if cell.value == name:
                                   self.worked = o
                                   Zenex_manager.sheet.cell(row=cell.row, column=7).value = o  
                                   Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
                                   Zenex_manager.wbk.close()
                elif selected == 5:
                    o = input('What would you like to replace this persons gender with?')
                    for row in Zenex_manager.sheet.iter_rows():
                        for cell in row:
                            if cell.value == name:
                                self.gender = o
                                Zenex_manager.sheet.cell(row=cell.row, column=9).value = str(o)
                                Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
                                Zenex_manager.wbk.close()
                                
                elif selected == 6:
                    o = input('What would you like to replace this persons name with?')
                    for row in Zenex_manager.sheet.iter_rows():
                        for cell in row:
                            if cell.value == name:
                                self.name = o
                                Zenex_manager.sheet.cell(row=cell.row, column=2).value = str(o)
                                Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
                                Zenex_manager.wbk.close()
                inputValid = True
                break
            else:
                print('Please select a valid ' + 'option' + ' number')
#------------------------ask you what you want to edit and then edits based on your input-----------------------------------------------------------



    def update_information_ID(self,ID=0):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      

#-------------------Checks if excel is open and whether it should close it or not------------------------       
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
#-------------------Checks if excel is open and whether it should close it or not------------------------       

#------checks if there is an employee here by inputted ID, whether it should use the selected employee or not, and as well as if the employee is a manager------------------------
        try:
            if ID == 0:
                ID = self.ID
        except AttributeError:
            raise Exception('you need to specify which employee you wanna update, either by doing emloyee.select_by_ID(ID) then employee.update_information_name() or employee.update_information_ID(ID)')
        file = pd.ExcelFile('Zenex worker data.xlsx')
        data = pd.read_excel(file, Zenex_manager.sheetvalue)
        if int(ID) not in data['employee ID'].values:
            raise Exception('There is no employee by that ID')
        data = data[data["employee ID"] == int(ID)]
        d = data['Is Manager'].values
        d = ' '.join(map(str,d))
        r = type(self)
        r = str(r)
        r = r = r.split('.')
        r = r[-1].replace("'", "")
        r = r.replace('>','')
        if d == 'Yes' or d == 'Executive': 
            if r == 'Zenex_Xmanager':
                if d == 'Yes':
                    pass
                elif d == 'Executive':
                    raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')                
            elif r == 'Zenex_CEO':
                pass
            else:
                raise Exception('This employee is a manager and or executive so you dont have athority to select or edit them')      
#------checks if there is an employee here by inputted ID, whether it should use the selected employee or not, and as well as if the employee is a manager------------------------
        
#------------------------ask you what you want to edit and then edits based on your input-----------------------------------------------------------
        options = {}
        options["Position"] = 1
        options["Race/ethnicity"] = 2
        options["Date employyed"] = 3
        options["Hours worked"] = 4
        options["Gender"] = 5
        options["Name"] = 6
        index = 0
        indexValidList = []
        print('which one do you want to edit?')
        for optionName in options:
            index = index + 1
            indexValidList.extend([options[optionName]])
            print(str(index) + ') ' + optionName)
        inputValid = False
        while not inputValid:
            inputRaw = input("option('must be a number'): ")
            inputNo = int(inputRaw) - 1
            if inputNo > -1 and inputNo < len(indexValidList):
                selected = indexValidList[inputNo]
                if selected == 1:
                    o = input('What would you like to replace this persons position with?')
                    rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
                    for row in reversed(rows):
                        if row[0].row == 1:
                            break
                        if int(row[0].value) == int(ID):
                            self.position = o
                            Zenex_manager.sheet.cell(row=row[0].row, column=4).value = o  
                            Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
                            Zenex_manager.wbk.close()
                elif selected == 2:
                    o = input('What would you like to replace this persons race with?')
                    rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
                    for row in reversed(rows):
                        if row[0].row == 1:
                            break
                        if int(row[0].value) == int(ID):
                            self.race = o
                            Zenex_manager.sheet.cell(row=row[0].row, column=5).value = o  
                            Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))
                            Zenex_manager.wbk.close()
                elif selected == 3:
                   try:
                       o = input('What would you like to replace this persons date employeed with?')
                       if (type(o) ==  datetime.date) == False:
                           o = parser.parse(o)
                           o = o.date()
                   except TypeError:
                       raise Exception('This is not a valid date')
                   rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
                   for row in reversed(rows):
                        if row[0].row == 1:
                            break
                        if int(row[0].value) == int(ID):
                            self.date_created = o
                            Zenex_manager.sheet.cell(row=row[0].row, column=6).value = o  
                            Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
                            Zenex_manager.wbk.close()
                elif selected == 4:
                    try:
                        o = input('What would you like to replace this persons hours worked with?')
                        o = int(o)
                    except ValueError:
                        raise Exception('{0} is not a valid number for this'.format(o))
                    rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
                    for row in reversed(rows):
                        if row[0].row == 1:
                            break
                        if int(row[0].value) == int(ID):
                            self.worked = o
                            Zenex_manager.sheet.cell(row=row[0].row, column=7).value = o  
                            Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
                            Zenex_manager.wbk.close()
                elif selected == 5:
                    o = input('What would you like to replace this persons gender with?')
                    rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
                    for row in reversed(rows):
                        if row[0].row == 1:
                            break
                        if int(row[0].value) == int(ID):
                            self.gender = o
                            Zenex_manager.sheet.cell(row=row[0].row, column=9).value = o  
                            Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx')) 
                            Zenex_manager.wbk.close()
                elif selected == 6:
                    o = input('What would you like to replace this persons name with?')
                    rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
                    for row in reversed(rows):
                        if row[0].row == 1:
                            break
                        if int(row[0].value) == int(ID):
                            self.name = o
                            Zenex_manager.sheet.cell(row=row[0].row, column=2).value = o  
                            Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
                            Zenex_manager.wbk.close()
                inputValid = True
                break
            else:
                print('Please select a valid ' + 'option' + ' number')
#------------------------ask you what you want to edit and then edits based on your input-----------------------------------------------------------



    def get_all_values(self,Type):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
        r = type(self)
        r = str(r)
        r = r = r.split('.')
        r = r[-1].replace("'", "")
        r = r.replace('>','')
        if r == 'Zenex_Xmanager' or 'Zenex_CEO':
            file = pd.ExcelFile('Zenex worker data.xlsx')
            data = pd.read_excel(file, Zenex_manager.sheetvalue)
        else:
            file = pd.ExcelFile('Zenex worker data.xlsx')
            data = pd.read_excel(file, Zenex_manager.sheetvalue)
            data = data[data['Is Manager'] != 'Yes']
            data = data[data['Is Manager'] != 'Executive']
                
        if Type.lower() == 'salary':
            return data['$ Salary per/hour $'].tolist()
        elif Type.lower() == 'name':
            return data['Name of employee'].tolist()
        elif Type.lower() == 'gender':
            return data['Gender'].tolist()
        elif Type.lower() == 'position':
            return data['Position of employee'].tolist()
        elif Type.lower() == 'increase':
            return data['Total Salary Increase'].tolist()
        elif Type.lower() == 'hours worked':
            return data['Hours Worked'].tolist()
        elif Type.lower() == 'dates':
            return data['Date employed'].tolist()
        elif Type.lower() == 'race':
            return data['Race/ethnicity'].tolist()
        elif Type.lower() == 'id':
            return data['employee ID'].tolist()
        else:
            raise Exception('That is not a valid choice, all inputs for this are salary,name,gender,position,increase,hours worked,dates,race, and ID')



    def __repr__(self):
        return "Zenex_manager('{}','{}')".format(self.sector,'Api key')
    
    def __str__(self):
        try:
            return 'Name: {}, Position: {}, Race: {}, Gender: {}, Salary: {}'.format(self.name, self.position, self.race, self.gender, self.salary) 
        except AttributeError:
            return 'Sector: {} (you have no selected employee at the moment'.format(self.sector)
            
            
    @classmethod
    def from_string(cls,string):
        sector,api_key = string.split('-')
        return cls(sector,api_key)
   
    def __gt__(self, other):
        options = {}
        options['which employee has a higher salary'] = 1
        options['which employee has been working here the longest'] = 2
        options['which employee has the highest salary increase'] = 3
        options['which employee has worked more hours'] = 4
        options['I want to use the normal python operators(using normal operators on an zenex_manager object is quite glitchy tho)'] = 5
        index = 0
        indexValidList = []
        print('Select a ' + 'option' + ':')
        for optionName in options:
            index = index + 1
            indexValidList.extend([options[optionName]])
            print(str(index) + ') ' + optionName)
        inputValid = False
        while not inputValid:
            inputRaw = input("option('must be a number'): ")
            inputNo = int(inputRaw) - 1
            if inputNo > -1 and inputNo < len(indexValidList):
                selected = indexValidList[inputNo]
                if selected == 1:
                    if hasattr(self,'salary') == True and hasattr(other,'salary') == True:
                        if self.salary > other.salary:
                            return True
                        else:
                            return False
                elif selected == 2:
                    if hasattr(self,'date_created') == True and hasattr(other,'date_created') == True:
                        if self.date_created > other.date_created:
                            return True
                        else:
                            return False
                elif selected == 3:
                    if hasattr(self,'increase') == True and hasattr(other,'increase') == True:
                        if self.increase > other.increase:
                            return True
                        else:
                            return False
                elif selected == 4:
                    if hasattr(self,'worked') == True and hasattr(other,'worked') == True:
                        if self.worked > other.worked:
                            return True
                        else:
                            return False
                elif selected == 5:
                    return NotImplemented
                inputValid = True
                break
            else:
                print('Please select a valid ' + 'option' + ' number')
                
                
                
    def __eq__(self, other):
        options = {}
        options['if the employees have the same salary'] = 1
        options['if the employees worked here for the exact amount of time'] = 2
        options['if the employees have the same salary increase'] = 3
        options['if the employees have worked the same hours'] = 4
        options['I want to use the normal python operators(using normal operators on an zenex_manager object is quite glitchy tho)'] = 5
        index = 0
        indexValidList = []
        print('Select a ' + 'option' + ':')
        for optionName in options:
            index = index + 1
            indexValidList.extend([options[optionName]])
            print(str(index) + ') ' + optionName)
        inputValid = False
        while not inputValid:
            inputRaw = input("option('must be a number'): ")
            inputNo = int(inputRaw) - 1
            if inputNo > -1 and inputNo < len(indexValidList):
                selected = indexValidList[inputNo]
                if selected == 1:
                    if hasattr(self,'salary') == True and hasattr(other,'salary') == True:
                        if self.salary == other.salary:
                            return True
                        else:
                            return False
                elif selected == 2:
                    if hasattr(self,'date_created') == True and hasattr(other,'date_created') == True:
                        if self.date_created == other.date_created:
                            return True
                        else:
                            return False
                elif selected == 3:
                    if hasattr(self,'increase') == True and hasattr(other,'increase') == True:
                        if self.increase == other.increase:
                            return True
                        else:
                            return False
                elif selected == 4:
                    if hasattr(self,'worked') == True and hasattr(other,'worked') == True:
                        if self.worked == other.worked:
                            return True
                        else:
                            return False
                elif selected == 5:
                    return NotImplemented
                    
                
                inputValid = True
                break
            else:
                print('Please select a valid ' + 'option' + ' number')
                
                
    def __ne__(self, other):
        options = {}
        options["if the employees don't'have the same salary"] = 1
        options['if the employees have not worked here for the exact amount of time'] = 2
        options['if the employees do not have the same salary increase'] = 3
        options["if the employees haven't worked the same hours"] = 4
        options['I want to use the normal python operators(using normal operators on an zenex_manager object is quite glitchy tho)'] = 5
        index = 0
        indexValidList = []
        print('Select a ' + 'option' + ':')
        for optionName in options:
            index = index + 1
            indexValidList.extend([options[optionName]])
            print(str(index) + ') ' + optionName)
        inputValid = False
        while not inputValid:
            inputRaw = input("option('must be a number'): ")
            inputNo = int(inputRaw) - 1
            if inputNo > -1 and inputNo < len(indexValidList):
                selected = indexValidList[inputNo]
                if selected == 1:
                    if hasattr(self,'salary') == True and hasattr(other,'salary') == True:
                        if self.salary != other.salary:
                            return True
                        else:
                            return False
                elif selected == 2:
                    if hasattr(self,'date_created') == True and hasattr(other,'date_created') == True:
                        if self.date_created != other.date_created:
                            return True
                        else:
                            return False
                elif selected == 3:
                    if hasattr(self,'increase') == True and hasattr(other,'increase') == True:
                        if self.increase != other.increase:
                            return True
                        else:
                            return False
                elif selected == 4:
                    if hasattr(self,'worked') == True and hasattr(other,'worked') == True:
                        if self.worked != other.worked:
                            return True
                        else:
                            return False
                elif selected == 5:
                    return NotImplemented
                
                inputValid = True
                break
            else:
                print('Please select a valid ' + 'option' + ' number')
                
    def __lt__(self, other):
        options = {}
        options['which employee has a lower salary'] = 1
        options['which employee has been working here the least'] = 2
        options['which employee has the lowest salary increase'] = 3
        options['which employee has worked less hours'] = 4
        options['I want to use the normal python operators(using normal operators on an zenex_manager object is quite glitchy tho)'] = 5
        index = 0
        indexValidList = []
        print('Select a ' + 'option' + ':')
        for optionName in options:
            index = index + 1
            indexValidList.extend([options[optionName]])
            print(str(index) + ') ' + optionName)
        inputValid = False
        while not inputValid:
            inputRaw = input("option('must be a number'): ")
            inputNo = int(inputRaw) - 1
            if inputNo > -1 and inputNo < len(indexValidList):
                selected = indexValidList[inputNo]
                if selected == 1:
                    if hasattr(self,'salary') == True and hasattr(other,'salary') == True:
                        if self.salary < other.salary:
                            return True
                        else:
                            return False
                elif selected == 2:
                    if hasattr(self,'date_created') == True and hasattr(other,'date_created') == True:
                        if self.date_created < other.date_created:
                            return True
                        else:
                            return False
                elif selected == 3:
                    if hasattr(self,'increase') == True and hasattr(other,'increase') == True:
                        if self.increase < other.increase:
                            return True
                        else:
                            return False
                elif selected == 4:
                    if hasattr(self,'worked') == True and hasattr(other,'worked') == True:
                        if self.worked < other.worked:
                            return True
                        else:
                            return False
                elif selected == 5:
                    return NotImplemented
                
                inputValid = True
                break
            else:
                print('Please select a valid ' + 'option' + ' number')
                
                
                
    def __le__(self, other):
        options = {}
        options["if {0}'s salary is less than or equal too {1}'s salary".format(self.name,other.name)] = 1
        options["if {0} has worked here for less time or equal time too {1}".format(self.name,other.name)] = 2
        options["if {0}'s salary increase is less than or equal too {1} salary increase".format(self.name,other.name)] = 3
        options["if {0} has worked less than or equal too {1}".format(self.name,other.name)] = 4
        options['I want to use the normal python operators(using normal operators on an zenex_manager object is quite glitchy tho)'] = 5
        index = 0
        indexValidList = []
        print('Select a ' + 'option' + ':')
        for optionName in options:
            index = index + 1
            indexValidList.extend([options[optionName]])
            print(str(index) + ') ' + optionName)
        inputValid = False
        while not inputValid:
            inputRaw = input("option('must be a number'): ")
            inputNo = int(inputRaw) - 1
            if inputNo > -1 and inputNo < len(indexValidList):
                selected = indexValidList[inputNo]
                if selected == 1:
                    if hasattr(self,'salary') == True and hasattr(other,'salary') == True:
                        if self.salary <= other.salary:
                            return True
                        else:
                            return False
                elif selected == 2:
                    if hasattr(self,'date_created') == True and hasattr(other,'date_created') == True:
                        if self.date_created <= other.date_created:
                            return True
                        else:
                            return False
                elif selected == 3:
                    if hasattr(self,'increase') == True and hasattr(other,'increase') == True:
                        if self.increase <= other.increase:
                            return True
                        else:
                            return False
                elif selected == 4:
                    if hasattr(self,'worked') == True and hasattr(other,'worked') == True:
                        if self.worked <= other.worked:
                            return True
                        else:
                            return False
                elif selected == 5:
                    return NotImplemented
                
                inputValid = True
                break
            else:
                print('Please select a valid ' + 'option' + ' number')
                
    def __ge__(self, other):
        options = {}
        options["if {0}'s salary is greater than or equal too {1}'s salary".format(self.name,other.name)] = 1
        options["if {0} has worked here for more time or equal time too {1}".format(self.name,other.name)] = 2
        options["if {0}'s salary increase is more than or equal too {1} salary increase".format(self.name,other.name)] = 3
        options["if {0} has worked more than or equal too {1}".format(self.name,other.name)] = 4
        options['I want to use the normal python operators(using normal operators on an zenex_manager object is quite glitchy tho)'] = 5
        index = 0
        indexValidList = []
        print('Select a ' + 'option' + ':')
        for optionName in options:
            index = index + 1
            indexValidList.extend([options[optionName]])
            print(str(index) + ') ' + optionName)
        inputValid = False
        while not inputValid:
            inputRaw = input("option('must be a number'): ")
            inputNo = int(inputRaw) - 1
            if inputNo > -1 and inputNo < len(indexValidList):
                selected = indexValidList[inputNo]
                if selected == 1:
                    if hasattr(self,'salary') == True and hasattr(other,'salary') == True:
                        if self.salary >= other.salary:
                            return True
                        else:
                            return False
                elif selected == 2:
                    if hasattr(self,'date_created') == True and hasattr(other,'date_created') == True:
                        if self.date_created >= other.date_created:
                            return True
                        else:
                            return False
                elif selected == 3:
                    if hasattr(self,'increase') == True and hasattr(other,'increase') == True:
                        if self.increase >= other.increase:
                            return True
                        else:
                            return False
                elif selected == 4:
                    if hasattr(self,'worked') == True and hasattr(other,'worked') == True:
                        if self.worked >= other.worked:
                            return True
                        else:
                            return False
                elif selected == 5:
                    return NotImplemented
                
                inputValid = True
                break
            else:
                print('Please select a valid ' + 'option' + ' number')


class Zenex_Xmanager(Zenex_manager):
    def __init__(self, sector, api_key,name,password):
        super().__init__(sector,api_key)
        things = []
        de = ''
        with open('zenames.txt') as f:
            r = f.read()
        r = r.split('\n')
        for i in r:
            w = i.split(' ',1)
            w = w[0]
            w = w.strip(':')
            if self.sector.capitalize() == w:
                r = i
        re = r.split(' ',1)
        re = re[1]
        re = re.split('.')
        for i in re:
            u = []
            Name = i.split(',')
            Password = Name[1]
            Name = Name[0]
            u.append(Name)
            u.append(Password)
            things.append(u)
        for i in things:
            if name == i[0] and password == i[1]:
                if password == 'Void':
                    de = input('You gotta change your password, what do you want it to be?')
                    de = str(de)
                break
        else:
            raise Exception("Either your name or password wasn't correct")
        with open('zenex_transfer.txt') as f:
            r = f.readlines()
        count = 0
        o = ''
        for i in r:
            if i != 'Zenex Transfer Sheet':   
                e = i.split(':')
                o = e[0].split(',')
            if 'Zenex Transfer Sheet' not in o:
                o = o[1]
            else:
                o=''
            if self.sector in o:
                count+= 1
        if count == 0:
            pass
        else:
            print('You have {0} new transfer request, {1}!'.format(count,name))
        self.Xname = name
        if de != '':
            self.change_my_info('password',de)
            
        
    def demote_name(self,name=''):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
        val = 'No'
        np = ''
        gp = ''
        yey = ''
        op = ''
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        file = pd.ExcelFile('Zenex worker data.xlsx')
        data = pd.read_excel(file, Zenex_manager.sheetvalue) 
        data = data[data['Name of employee'] == str(name)]
        d = data['Is Manager'].values
        d = ' '.join(map(str,d))
        r = type(self)
        r = str(r)
        r = r = r.split('.')
        r = r[-1].replace("'", "")
        r = r.replace('>','')
        if d == 'Executive': 
            if r == 'Zenex_CEO':
                g = False
                while not g:
                    shall = input('This employee is a Executive, do you want to demote him to manager or employee(ouch)?')
                    if shall.lower() == 'manager':
                        val = 'Yes'
                        yey = 'yes'
                        op = data['Name of employee'].values
                        op = ' '.join(map(str,op))
                        np = 36
                        gp = (np - int(data['$ Salary per/hour $'].values)) + int(data['Total Salary Increase'].values)
                        g = True
                    elif shall.lower() == 'employee':
                        val = 'No'
                        yey = 'yes'
                        op = data['Name of employee'].values
                        op = ' '.join(map(str,op))
                        np = 36
                        gp = (np - int(data['$ Salary per/hour $'].values)) + int(data['Total Salary Increase'].values)
                        g = True
                    else:
                        print('Not a valid answer, either put employee or manager')
            else:
                raise Exception('This employee is a executive so you dont have athority to select or edit them')
#-------------------Checks if excel is open and whether it should close it or not------------------------       
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
#-------------------Checks if excel is open and whether it should close it or not------------------------       

#------checks if there is an employee here by inputted name, whether it should use the selected employee or not, whether there are 2 employees by 1 name------------------------
        try:
            if name == '':
                name = self.name
        except AttributeError:
            raise Exception('you need to specify which employee you wanna demote, either by doing emloyee.select_by_name(name) then employee.demote_name() or employee.demote_name(name)')
        if name not in data['Name of employee'].values:
            raise Exception('There is no employee by the name {0} (this is case sensitive so make sure your typing it in right)'.format(name))
        if len(data) != 1:
            raise Exception('there are 2 employee by the name {0}, try searching by employee ID'.format(name))
#------checks if there is an employee here by inputted name, whether it should use the selected employee or not, whether there are 2 employees by 1 name------------------------

        for row in Zenex_manager.sheet.iter_rows():
            for cell in row:
                if cell.value == name:
                    Zenex_manager.sheet.cell(row=cell.row, column=10).value = val
                    if np != '':
                        lio = []
                        ew = 1
                        sd = 0
                        Zenex_manager.sheet.cell(row=cell.row, column=3).value = int(np)
                        Zenex_manager.sheet.cell(row=cell.row, column=8).value = int(gp)
                        Zenex_manager.sheet.insert_rows(Zenex_manager.sheet.max_row)
                        for i in range(1,11):
                            lio.append(Zenex_manager.sheet.cell(row=cell.row, column=i).value)
                        rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
                        for row in reversed(rows):
                            #print(row[0].value)
                            if row[1].row == 1:
                                break
                            if str(row[1].value) == str(name):
                                Zenex_manager.sheet.delete_rows(row[1].row, 1)
                        for row in Zenex_manager.sheet.iter_rows():
                            for cell in row:
                                if cell.value == None:
                                    Zenex_manager.sheet.cell(row=cell.row, column=ew).value = lio[sd]
                                    ew+=1
                                    sd+=1
        rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
        for row in reversed(rows):
            if row[0].row == 1:
                break
            if row[0].value == None:
                Zenex_manager.sheet.delete_rows(row[0].row, 1)
        
        Zenex_manager.wbk.save(filename=os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
        Zenex_manager.wbk.close()
        if yey == 'yes':
            with open('zenames.txt') as f:
                yu = f.readlines()
            with open('zenames.txt','w') as f:
                for i in yu:
                    if self.sector.capitalize() in i:
                        i = i.strip(' '+self.sector.capitalize()+':')
                        i = i.split('.')
                        for w in i:
                            if op in w:
                                r = i.index(w)
                                del i[r]
                        fr = ''
                        for te in i:
                            if te != i[-1]:
                                fr = fr + te + '.'
                            else:
                                fr = fr + te
                        fr = self.sector.capitalize() + ': ' + fr
                        if '.' not in fr:
                            fr = fr + '\n'
                        f.write(fr)
                    else:
                        f.write(i.strip('\n')+'\n')
        
        try:
            if name == self.name:
                del self.is_manager
                del self.ID
                del self.name
                del self.race
                del self.salary
                del self.position
                del self.increase
                del self.gender
                del self.worked
                del self.date_created
        except AttributeError:
            pass
        
    def demote_ID(self,ID=0):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
        val = 'No'
        np = ''
        yey = ''
        gp = ''
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      

#-------------------Checks if excel is open and whether it should close it or not------------------------       
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
#-------------------Checks if excel is open and whether it should close it or not------------------------       

#------checks if there is an employee here by inputted ID, whether it should use the selected employee or not, and as well as if the employee is a manager------------------------
        try:
            if ID == 0:
                ID = self.ID
        except AttributeError:
            raise Exception('you need to specify which employee you wanna demote, either by doing emloyee.select_by_ID(ID) then employee.demote_ID() or employee.demote_ID(ID)')
        file = pd.ExcelFile('Zenex worker data.xlsx')
        data = pd.read_excel(file, Zenex_manager.sheetvalue)
        if int(ID) not in data['employee ID'].values:
            raise Exception('There is no employee by that ID')
        data = data[data["employee ID"] == int(ID)]
        d = data['Is Manager'].values
        d = ' '.join(map(str,d))
        r = type(self)
        r = str(r)
        r = r = r.split('.')
        r = r[-1].replace("'", "")
        r = r.replace('>','')
        if d == 'Executive': 
            if r == 'Zenex_CEO':
                g = False
                while not g:
                    shall = input('This employee is a Executive, do you want to demote him to manager or employee(ouch)?')
                    if shall.lower() == 'manager':
                        val = 'Yes'
                        yey = 'yes'
                        op = data['Name of employee'].values
                        op = ' '.join(map(str,op))
                        np = 36
                        gp = (np - int(data['$ Salary per/hour $'].values)) + int(data['Total Salary Increase'].values)
                        g = True
                    elif shall.lower() == 'employee':
                        val = 'No'
                        yey = 'yes'
                        op = data['Name of employee'].values
                        op = ' '.join(map(str,op))
                        np = 36
                        gp = (np - int(data['$ Salary per/hour $'].values)) + int(data['Total Salary Increase'].values)
                        g = True
                    else:
                        print('Not a valid answer, either put employee or manager')
            else:
                raise Exception('This employee is a executive so you dont have athority to select or edit them')
                
#------checks if there is an employee here by inputted ID, whether it should use the selected employee or not, and as well as if the employee is a manager------------------------

        for row in Zenex_manager.sheet.iter_rows():
            for cell in row:
                if cell.value == ID:
                    Zenex_manager.sheet.cell(row=cell.row, column=10).value = val
                    if np != '':
                        lio = []
                        ew = 1
                        sd = 0
                        Zenex_manager.sheet.cell(row=cell.row, column=3).value = int(np)
                        Zenex_manager.sheet.cell(row=cell.row, column=8).value = int(gp)
                        Zenex_manager.sheet.insert_rows(Zenex_manager.sheet.max_row)
                        for i in range(1,11):
                            lio.append(Zenex_manager.sheet.cell(row=cell.row, column=i).value)
                        rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
                        for row in reversed(rows):
                            if row[0].row == 1:
                                break
                            try:
                                if int(row[0].value) == int(ID):
                                    Zenex_manager.sheet.delete_rows(row[1].row, 1)
                            except TypeError:
                                pass
                        for row in Zenex_manager.sheet.iter_rows():
                            for cell in row:
                                if cell.value == None:
                                    Zenex_manager.sheet.cell(row=cell.row, column=ew).value = lio[sd]
                                    ew+=1
                                    sd+=1
        rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
        for row in reversed(rows):
            if row[0].row == 1:
                break
            if row[0].value == None:
                Zenex_manager.sheet.delete_rows(row[0].row, 1)
        try:
            if ID == self.ID:
                del self.is_manager
                del self.ID
                del self.name
                del self.race
                del self.salary
                del self.position
                del self.increase
                del self.gender
                del self.worked
                del self.date_created
        except AttributeError:
            pass
        Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))    
        Zenex_manager.wbk.close()
        if yey == 'yes':
            with open('zenames.txt') as f:
                yu = f.readlines()
            with open('zenames.txt','w') as f:
                for i in yu:
                    if self.sector.capitalize() in i:
                        i = i.strip(' '+self.sector.capitalize()+':')
                        i = i.split('.')
                        for w in i:
                            if op in w:
                                r = i.index(w)
                                del i[r]
                        fr = ''
                        for te in i:
                            if te != i[-1]:
                                fr = fr + te + '.'
                            else:
                                fr = fr + te
                        fr = self.sector.capitalize() + ': ' + fr
                        if '.' not in fr:
                            fr = fr + '\n'
                        f.write(fr)
                    else:
                        f.write(i.strip('\n')+'\n')
        
        
        
    def transfer_employee(self,sector,ID=0):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
        try:
            if ID == 0:
                ID = self.ID
        except AttributeError:
            raise Exception('you need to specify which employee you wanna transfer, either by doing emloyee.select_by_ID(ID) then employee.request_transfer() or employee.request_transfer(ID)')
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy 
        if sector == '':
            raise Exception('you need to put which job sector your in! failed to load sector information')
        elif sector.lower() not in Zenex_manager.sectors:
            raise Exception('{0} is not a job sector here! failed to load sector information'.format(sector))
        elif sector.lower() == self.sector:
            raise Exception("Can't transfer to your own sector!")

#-------------------Checks if excel is open and whether it should close it or not------------------------       
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
#-------------------Checks if excel is open and whether it should close it or not------------------------       
        with open('zenex_transfer.txt') as f:
            r = f.read()
        r = r.split('\n')
        tyu = ''
        for i in r:
            if i != 'Zenex Transfer Sheet:':   
                tyu = i.split(',')
            if len(tyu) != 0:
                tr = tyu[1]
            else:
                tr = ''
            if str(ID) in tr:
                raise Exception('employee is already up for transfer')
                
#------checks if there is an employee here by inputted ID, and whether it should use the selected employee or not------------------------
        file = pd.ExcelFile('Zenex worker data.xlsx')
        data = pd.read_excel(file, Zenex_manager.sheetvalue)
        if int(ID) not in data['employee ID'].values:
            raise Exception('There is no employee by that ID')
        data = data[data["employee ID"] == int(ID)]
        d = data['Is Manager'].values
        d = ' '.join(map(str,d))
        r = type(self)
        r = str(r)
        r = r = r.split('.')
        r = r[-1].replace("'", "")
        r = r.replace('>','')
        if d == 'Executive': 
            if r == 'Zenex_CEO':
                pass
            else:
                raise Exception('This employee is a executive so you dont have athority to select or edit them')
                
#------checks if there is an employee here by inputted ID, and whether it should use the selected employee or not------------------------
        self.is_manager = data['Is Manager'].values
        self.is_manager = ' '.join(map(str,self.is_manager))
        self.ID = data['employee ID'].values
        self.ID = ' '.join(map(str,self.ID))
        self.name = data['Name of employee'].values
        self.name = ' '.join(map(str,self.name))
        self.race = data['Race/ethnicity'].values
        self.race = ' '.join(map(str,self.race))
        self.salary = data['$ Salary per/hour $'].values
        self.salary = ' '.join(map(str,self.salary))
        self.salary = int(self.salary)
        self.position = data['Position of employee'].values
        self.position = ' '.join(map(str,self.position))
        self.increase = data['Total Salary Increase'].values
        self.increase = ' '.join(map(str,self.increase))
        self.increase = int(self.increase)
        self.gender = data['Gender'].values
        self.gender = ' '.join(map(str,self.gender))
        self.worked = data['Hours Worked'].values
        self.worked = ' '.join(map(str,self.worked))
        self.worked = int(self.worked)
        self.date_created = data['Date employed'].values
        self.date_created = ' '.join(map(str,self.date_created))
        self.date_created = parser.parse(self.date_created)
        self.date_created = self.date_created.date() 
        values = [self.ID,self.name,self.salary,self.position,self.race,self.date_created,self.worked,self.increase,self.gender]
        with open('zenex_transfer.txt','a') as f:
            f.write('\n{0},transfer request to the {1} sector: '.format(self.sector,sector))
            for r in values:
                f.write(str(r)+',')
            f.write(str(self.is_manager))
        try:
            if ID == self.ID:
                del self.is_manager
                del self.ID
                del self.name
                del self.race
                del self.salary
                del self.position
                del self.increase
                del self.gender
                del self.worked
                del self.date_created
        except AttributeError:
            pass
            
    def manage_request(self):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy

#-------------------Checks if excel is open and whether it should close it or not------------------------       
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
#-------------------Checks if excel is open and whether it should close it or not------------------------       
        with open('zenex_transfer.txt') as f:
            r = f.readlines()
        request = []
        for i in r:
            if i != 'Zenex Transfer Sheet':   
                e = i.split(':')
                o = e[0].split(',')
            if 'Zenex Transfer Sheet' not in o:
                o = o[1]
            else:
                o=''
            if self.sector in o:
                request.append(i)
        options = {}
        for i in request:
            options[i] = i
        if options:
            pass
        else:
            print('You have no request :)')
            sys.exit()
        index = 0
        indexValidList = []
        print('which one do you want to reveiw?')
        for optionName in options:
            index = index + 1
            indexValidList.extend([options[optionName]])
            print(str(index) + ') ' + optionName)
        inputValid = False
        while not inputValid:
            inputRaw = input("option('must be a number'): ")
            inputNo = int(inputRaw) - 1
            if inputNo > -1 and inputNo < len(indexValidList):
                selected = indexValidList[inputNo]
                for i in request:
                    if selected == i:
                        q = False
                        print('request:',i)
                        while not q:
                            p = input('Will you accept this request?(Yes or No)')
                            if p.lower() == 'yes':
                                print('accepted!')
                                y = i.split(':')
                                y = y[1].split(' ',1)
                                y = y[1].split(',')
                                op = False
                                while not op:
                                    gh = False
                                    ju = y[2]
                                    oi = y[7]
                                    while not gh:
                                        answer = input('do you want to increase or decrease salary?')
                                        if answer.lower() == 'increase':
                                            uy = input('Input how much you want to increase salary(salary = {0}):'.format(y[2]))
                                            if y[7] == 0:
                                                y[7] = uy
                                            else:
                                                y[7] = int(y[7]) + int(uy)
                                            y[2] = int(y[2]) + int(uy)
                                            gh = True
                                        elif answer.lower() == 'decrease':
                                            uy = input('Input how much you want to decrease salary(salary = {0}):'.format(y[2]))
                                            if y[7] == 0:
                                                y[7] = uy
                                            else:
                                                y[7] = int(y[7]) - int(uy)
                                            y[2] = int(y[2]) - int(uy)
                                            gh = True
                                        else:
                                            print('answer must be yes or no')
                                    if int(y[2]) > 25 or int(y[2]) < 15:
                                        print("{0} isn't a valid salary for this company".format(str(y[2])))
                                        y[2] = ju
                                        y[7] = oi
                                        uy = 0
                                    else:
                                        op = True
                                y[3] = input('Input new position:')
                                row = 2
                                column = 1
                                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
                                while Zenex_manager.sheet.cell(row = row, column=column).value != None:
                                    row = row + 1
                                for i in y:
                                    if i != y[5]:
                                        try:
                                            Zenex_manager.sheet.cell(row=row, column=column).value = int(i)
                                            column = column + 1
                                        except ValueError:
                                            Zenex_manager.sheet.cell(row=row, column=column).value = i.strip('\n')
                                            column = column + 1
                                    elif i == y[5]:
                                        d = parser.parse(i)
                                        d = d.date()
                                        Zenex_manager.sheet.cell(row=row, column=column).value = d
                                        column = column + 1
                                request = []
                                for i in r:
                                    e = i.split(':')
                                    request.append(i)
                                with open("zenex_transfer.txt", "w") as f:
                                    f.write(request[0].strip('\n'))
                                    request.remove(request[0])
                                    for i in request:
                                        if i != selected:
                                            f.write('\n'+i.strip('\n'))
                                with open('zenex_transfer.txt') as f:
                                    r = f.read()
                                r = r.split('\n')
                                ty = selected.split(',')
                                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(ty[0])
                                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
                                rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
                                for row in reversed(rows):
                                    if row[0].row == 1:
                                        break
                                    if int(row[0].value) == int(y[0]):
                                        Zenex_manager.sheet.delete_rows(row[0].row, 1)
                                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
                                    
                                q = True
                            elif p.lower() == 'no':
                                print('Declined')
                                request = []
                                for i in r:
                                    e = i.split(':')
                                    request.append(i)
                                with open("zenex_transfer.txt", "w") as f:
                                    f.write(request[0].strip('\n'))
                                    request.remove(request[0])
                                    for i in request:
                                        if i != selected:
                                            f.write('\n'+i.strip('\n'))
                                q = True
                            else:
                                print('Answer must be Yes or No')
                inputValid = True
                break
            else:
                print('Please select a valid option number')
        Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
        Zenex_manager.wbk.close()
        
        
        
        
    def change_my_info(self,thingtochange,newvalue):
        if thingtochange == 'name':
            if '.' in newvalue or ',' in newvalue:
                raise Exception('No periods or commas allowed in naming!')
        with open('zenames.txt') as f:
            yu = f.readlines()
        with open('zenames.txt','w') as f:
            for i in yu:
                if self.sector.capitalize() in i:
                    i = i.replace(self.sector.capitalize()+': ','')
                    i = i.split('.')
                    for w in i:
                        if self.Xname in w:
                            r = i.index(w)
                            w = w.split(',')
                            e = ''
                            if thingtochange.lower() == 'password':
                                w[1] = str(newvalue)
                                for u in w:
                                    if u != w[-1]:
                                        e = e + u + ','
                                    else:
                                        e = e+u
                                i[r] = e
                            elif thingtochange.lower() == 'name':
                                w[0] = str(newvalue)
                                for u in w:
                                    if u != w[-1]:
                                        e = e + u + ','
                                    else:
                                        e = e+u
                                i[r] = e
                            else:
                                raise Exception("{0} is not a valid option, either put 'password' or 'name'".format(thingtochange))
                    fr = ''
                    for te in i:
                        if te != i[-1]:
                            fr = fr + te + '.'
                        else:
                            fr = fr + te
                    fr = self.sector.capitalize() + ': ' + fr
                    if '.' not in fr:
                        fr = fr + '\n'
                    if fr[-2] + fr[-1] != '\n':
                        fr = fr + '\n'
                    f.write(fr.strip('\n')+'\n')
                else:
                    f.write(i.strip('\n')+'\n')
                


class Zenex_CEO(Zenex_Xmanager):
    def __init__(self,name,password,sector):
        if sector == '':
            raise Exception('you need to put which job sector your in! failed to load sector information')
        if sector.lower() in Zenex_manager.sectors:
            self.sector = sector.lower()
            Zenex_manager.sheetvalue = ('Zenex {0}_sector'.format(self.sector))
            Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
        else:
            raise Exception('{0} is not a job sector here! failed to load sector information'.format(sector))
        with open('ZEO.txt') as f:
            r = f.read() 
        r = r.split(',')
        if name.lower() != r[0].lower():
            raise Exception("i dont know who {0} is, but it isn't Paul!".format(name))
        if password != r[1]:
            raise Exception("{0} isn't the password".format(password))
        cons = ['What is your favorite food?','Which ones do you like better? Humans or digital things like games?','Are you fat?','Are you relaxed right now?']
        question = random.choice(cons)
        answer = input(question)
        with open('ansy.txt') as fw:
            fw = fw.read()
        fw = fw.split('\n')
        if question == 'What is your favorite food?':
            if answer.lower() != fw[0]:
                raise Exception('Wrong!')
            else:
                print('They really are the best!')
        elif question == 'Which ones do you like better? Humans or digital things like games?':
            if answer.lower() != fw[1]:
                raise Exception('Wrong!')
            else:
                print('Good, your still you!')
        elif question == 'Are you fat?':
            if answer.lower() != fw[2]:
                raise Exception('Wrong!')
            else:
                print('hard truth')
        elif question == 'Are you relaxed right now?':
            if answer.lower() != fw[3]:
                raise Exception('Wrong!')
            else:
                print('sucks i know')
                
    def change_sector(self,sector):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
        if sector == '':
            raise Exception('you need to put which job sector your in! failed to load sector information')
        if sector.lower() in Zenex_manager.sectors:
            self.sector = sector.lower()
            Zenex_manager.sheetvalue = ('Zenex {0}_sector'.format(self.sector))
            Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
        else:
            raise Exception('{0} is not a job sector here! failed to load sector information'.format(sector))
        try:
            del self.is_manager
            del self.ID
            del self.name
            del self.race
            del self.salary
            del self.position
            del self.increase
            del self.gender
            del self.worked
            del self.date_created
        except AttributeError:
            pass
        
        
        
    def transfer_employee(self,inital_sector,new_sector,ID=0):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
        try:
            if ID == 0:
                ID = self.ID
        except AttributeError:
            raise Exception('you need to specify which employee you wanna transfer, either by doing emloyee.select_by_ID(ID) then employee.request_transfer() or employee.request_transfer(ID)')
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy      
        for s in Zenex_manager.sectors:
            if inital_sector.lower() == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(inital_sector.lower())
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
#Checks whther you working on the correct sheet in the file(so people using multiple objects can rest easy 
        if inital_sector == '' or new_sector == '':
            raise Exception('you need to put which job sector your transfer to and which your transfering from! in! failed to load sector information')
        elif inital_sector.lower() not in Zenex_manager.sectors or new_sector.lower() not in Zenex_manager.sectors:
            raise Exception('{0} is not a job sector here! failed to load sector information(come on you the ceo, you should no this)'.format(inital_sector.lower()))
        elif inital_sector.lower() == new_sector.lower():
            raise Exception('The guy is already in the {0} sector'.format(inital_sector.lower()))

#-------------------Checks if excel is open and whether it should close it or not------------------------       
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
            
        with open('zenex_transfer.txt') as f:
            r = f.read()
        r = r.split('\n')
        tyu = ''
        for i in r:
            if i != 'Zenex Transfer Sheet:':   
                tyu = i.split(',')
            if len(tyu) != 0:
                tr = tyu[1]
            else:
                tr = ''
            if str(ID) in tr:
                raise Exception('employee is already up for transfer')
        file = pd.ExcelFile('Zenex worker data.xlsx')
        data = pd.read_excel(file, Zenex_manager.sheetvalue)
        if int(ID) not in data['employee ID'].values:
            raise Exception('There is no employee by that ID')
        data = data[data["employee ID"] == int(ID)]
                
#------checks if there is an employee here by inputted ID, and whether it should use the selected employee or not------------------------
        self.is_manager = data['Is Manager'].values
        self.is_manager = ' '.join(map(str,self.is_manager))
        self.ID = data['employee ID'].values
        self.ID = ' '.join(map(str,self.ID))
        self.name = data['Name of employee'].values
        self.name = ' '.join(map(str,self.name))
        self.race = data['Race/ethnicity'].values
        self.race = ' '.join(map(str,self.race))
        self.salary = data['$ Salary per/hour $'].values
        self.salary = ' '.join(map(str,self.salary))
        self.salary = int(self.salary)
        self.position = data['Position of employee'].values
        self.position = ' '.join(map(str,self.position))
        self.increase = data['Total Salary Increase'].values
        self.increase = ' '.join(map(str,self.increase))
        self.increase = int(self.increase)
        self.gender = data['Gender'].values
        self.gender = ' '.join(map(str,self.gender))
        self.worked = data['Hours Worked'].values
        self.worked = ' '.join(map(str,self.worked))
        self.worked = int(self.worked)
        self.date_created = data['Date employed'].values
        self.date_created = ' '.join(map(str,self.date_created))
        self.date_created = parser.parse(self.date_created)
        self.date_created = self.date_created.date() 
        values = [self.ID,self.name,self.salary,self.position,self.race,self.date_created,self.worked,self.increase,self.gender,self.is_manager]
        ju = values[2]
        oi = values[7]
        op = False
        while not op:
            gh = False
            while not gh:
                answer = input('do you want to increase or decrease salary?')
                if answer.lower() == 'increase':
                    uy = input('Input how much you want to increase salary(salary = {0}):'.format(values[2]))
                    if values[7] == 0:
                        values[7] = uy
                    else:
                        values[7] = int(values[7]) + int(uy)
                        values[2] = int(values[2]) + int(uy)
                    gh = True
                elif answer.lower() == 'decrease':
                    uy = input('Input how much you want to decrease salary(salary = {0}):'.format(values[2]))
                    if values[7] == 0:
                        values[7] = uy
                    else:
                        values[7] = int(values[7]) - int(uy)
                    values[2] = int(values[2]) - int(uy)
                    gh = True
                else:
                    print('answer must be yes or no')
            if values[9] == 'Executive':
                if int(values[2]) > 100 or int(values[2]) < 60:
                    print("{0} isn't a valid salary for a company Executive".format(str(values[2])))
                    values[2] = ju
                    values[7] = oi
                    uy = 0
                else:
                    op = True
            else:    
                if int(values[2]) > 25 or int(values[2]) < 15:
                    print("{0} isn't a valid salary for this company".format(str(values[2])))
                    values[2] = ju
                    values[7] = oi
                    uy = 0
                else:
                    op = True
        values[3] = input('Input new position:')
        row = 2
        column = 1
        Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(new_sector.lower())
        Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
        if values[9] == 'Executive':
            Zenex_manager.sheet.insert_rows(2)
        while Zenex_manager.sheet.cell(row = row, column=column).value != None:
            row = row + 1
        for i in values:
            if i != values[5]:
                try:
                    Zenex_manager.sheet.cell(row=row, column=column).value = int(i)
                    column = column + 1
                except ValueError:
                    Zenex_manager.sheet.cell(row=row, column=column).value = i.strip('\n')
                    column = column + 1
            elif i == values[5]:
                Zenex_manager.sheet.cell(row=row, column=column).value = i
                column = column + 1
        Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(inital_sector.lower())
        Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
        rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
        for row in reversed(rows):
            if row[0].row == 1:
                break
            if int(row[0].value) == int(values[0]):
                Zenex_manager.sheet.delete_rows(row[0].row, 1)
        try:
            if ID == self.ID:
                del self.is_manager
                del self.ID
                del self.name
                del self.race
                del self.salary
                del self.position
                del self.increase
                del self.gender
                del self.worked
                del self.date_created
                
        except AttributeError:
            pass
        Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
        Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
        Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
        Zenex_manager.wbk.close()
        
        
        
    def change_my_info(self,thingtochange,newvalue):
        if thingtochange.lower() == 'question answer':
            with open('ansy.txt') as fw:
                fw = fw.read()
            fw = fw.split('\n')
            cons = ['What is your favorite food?','Which ones do you like better? Humans or digital things like games?','Are you fat?','Are you relaxed right now?']
            f = ['','','','']
            for i in range(4):
                f[i] = cons[i]+': ' + fw[i]
            options = {}
            for i in f:
                options[i] = i
            index = 0
            indexValidList = []
            print('which one do you want to reveiw?')
            for optionName in options:
                index = index + 1
                indexValidList.extend([options[optionName]])
                print(str(index) + ') ' + optionName)
            inputValid = False
            while not inputValid:
                inputRaw = input("option('must be a number'): ")
                inputNo = int(inputRaw) - 1
                if inputNo > -1 and inputNo < len(indexValidList):
                    selected = indexValidList[inputNo]
                    for i in f:
                        if selected == i:
                            fw[f.index(i)] = newvalue
                            inputValid = True
            for i in fw:
                fw[fw.index(i)] = fw[fw.index(i)]+'\n'
            fw[-1] = fw[-1].strip('\n')
            with open('ansy.txt','w') as f:
                for i in fw:
                    f.write(i)
        else:
            with open('ZEO.txt') as f:
                ZEO = f.read()
            ZEO = ZEO.split(',')
            if thingtochange.lower() == 'password':
                ZEO[1] = newvalue
            else:
                raise Exception("{0} is not a valid option, either put 'password' or 'question answer'".format(thingtochange))
            ZEO[0] = ZEO[0] + ','
            with open('ZEO.txt','w') as f:
                for i in ZEO:
                    f.write(i)
    
    def clear_sector(self,sector=''):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
        if sector == '':
            for s in Zenex_manager.sectors:
                if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                    Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                    Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
                    sector = self.sector
        else:
            if sector.lower() in Zenex_manager.sectors:
                Zenex_manager.sheetvalue = ('Zenex {0}_sector'.format(sector.lower()))
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
            elif sector.lower() == 'all':
                pass
            else:
                raise Exception('{0} is not a job sector here'.format(sector))
        w = False
        while not w:
            ig = input('Are you sure? you cant undo this:')
            if ig.lower() == 'yes':
                if sector.lower() == 'all':
                    for s in Zenex_manager.sectors:
                        Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(s)
                        Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
                        rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
                        for row in reversed(rows):
                            if row[9].row == 1:
                                break
                            if row[9].value != 'Executive':
                                Zenex_manager.sheet.delete_rows(row[0].row, 1)
                        w = True
                else:
                    rows = list(Zenex_manager.sheet.iter_rows(min_row=1, max_row=Zenex_manager.sheet.max_row))
                    for row in reversed(rows):
                        if row[9].row == 1:
                            break
                        if row[9].value != 'Executive':
                            Zenex_manager.sheet.delete_rows(row[0].row, 1)
                            w = True
            elif ig.lower() == 'no':
                w = True
            else:
                print('Not a valid answer, either yes or no')        
        Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
        Zenex_manager.wbk.close()
        for s in Zenex_manager.sectors:
            if self.sector == s and Zenex_manager.sheetvalue != 'Zenex {0}_sector'.format(s):
                Zenex_manager.sheetvalue = 'Zenex {0}_sector'.format(self.sector)
                Zenex_manager.sheet = Zenex_manager.wbk.get_sheet_by_name(Zenex_manager.sheetvalue)
        try:
            if sector.lower() == self.sector or sector.lower() == 'all':
                del self.is_manager
                del self.ID
                del self.name
                del self.race
                del self.salary
                del self.position
                del self.increase
                del self.gender
                del self.worked
                del self.date_created      
        except AttributeError:
            pass
        
                
                
    def create_sector(self,sector_name):
        if ',' in sector_name or '.' in sector_name:
            raise Exception('No periods or commas allowed in naming!')
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
        if sector_name == '':
            raise Exception('you need to put which job sector your in! failed to load sector information')
        if sector_name.lower() in Zenex_manager.sectors:
            raise Exception('{0} is already a sector here!'.format(sector_name))
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
        std = 'Zenex {0}_sector'.format(sector_name)
        Zenex_manager.wbk.create_sheet(std)
        Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
        Zenex_manager.wbk.close()
        with open('sec.txt','a') as f:
            f.write(','+sector_name.lower())
        characters = list(string.ascii_letters + string.digits + "!@#$%^&*()")
        random.shuffle(characters)
        password = []
        for i in range(30):
             password.append(random.choice(characters))
        random.shuffle(password)
        code = "".join(password)
        for i in range(8):
             password.append(random.choice(characters))
        with open('zenapi.txt','a') as f:
            f.write('\n'+sector_name.capitalize() + ': ' + code)
        print('You gotta create a exective for this sector')
        gfre = False
        while not gfre:
            name = input('Name?')
            name = str(name)
            position = input('position?')
            position = str(position)
            race = input('race?')
            race = str(race)
            gender = input('gender?')
            gender = str(gender)
            if ',' in name or '.' in name:
                print('No periods or commas allowed in naming!')
            elif ',' in position or '.' in position:
                print('No periods or commas allowed in naming!')
            elif ',' in race or '.' in race:
                print('No periods or commas allowed in naming!')
            elif ',' in gender or '.' in gender:
                print('No periods or commas allowed in naming!')
            else:
                gfre = True
        g = False
        while not g:
            salary = input('salary?')
            try:
                salary = int(salary)
                if salary > 100 or salary < 60:
                    print("{0} isn't a valid salary for this company".format(str(salary)))
                else:
                    g = True
            except ValueError:
                print('{0} is not a number'.format(str(salary)))
        Zenex_manager.sectors.append(sector_name.lower())
        self.change_sector(sector_name)
        e = ['employee ID', 'Name of employee', '$ Salary per/hour $',
             'Position of employee', 'Race/ethnicity', 'Date employed',
             'Hours Worked', 'Total Salary Increase', 'Gender', 'Is Manager']
        qwr = 1
        for i in e:
            Zenex_manager.sheet.cell(row=1, column=qwr).value = i
            qwr+=1
        Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
        Zenex_manager.wbk.close()
        passworde = ''.join((random.choice('123456789') for i in range(8)))
        d = [int(passworde),name,salary,position,race,datetime.datetime.now().date(),0,0,gender,'Executive']
        LK = 1
        for i in d:
            Zenex_manager.sheet.cell(row=2, column=LK).value = i
            LK+=1
        Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx')) 
        Zenex_manager.wbk.close()
        with open('zenames.txt','a') as f:
            f.write('\n'+sector_name.capitalize() + ': '+name+',Void')
        try:
            del self.is_manager
            del self.ID
            del self.name
            del self.race
            del self.salary
            del self.position
            del self.increase
            del self.gender
            del self.worked
            del self.date_created
        except AttributeError:
            pass
    def remove_sector(self,sector_name):
        with open('sec.txt') as f:
            Zenex_manager.sectors = f.read()
        Zenex_manager.sectors = Zenex_manager.sectors.split(',')  
        if sector_name == '':
            raise Exception('you need to put which job sector your in! failed to load sector information')
        if sector_name.lower() not in Zenex_manager.sectors:
            raise Exception('{0} is not a sector here!'.format(sector_name))
        if "EXCEL.EXE" in (i.name() for i in psutil.process_iter()):
            subprocess.run('TASKKILL /F /IM Excel.exe',capture_output=True)
            t.sleep(1)
        std = Zenex_manager.wbk.get_sheet_by_name('Zenex {0}_sector'.format(sector_name))
        Zenex_manager.wbk.remove_sheet(std)
        Zenex_manager.wbk.save(os.path.join(os.getcwd(),'Zenex worker data.xlsx'))  
        Zenex_manager.wbk.close()
        with open('zenames.txt') as f:
            yu = f.readlines()
        for i in yu:
            if sector_name.capitalize() in i:
                del yu[yu.index(i)]
        yu[-1] = yu[-1].strip('\n')
        with open('zenames.txt','w') as f:
            for i in yu:
                f.write(i)
        with open('zenapi.txt') as f:
            yua = f.readlines()
        for i in yua:
            if sector_name.capitalize() in i:
                del yua[yua.index(i)]
        yua[-1] = yua[-1].strip('\n')
        with open('zenapi.txt','w') as f:
            for i in yua:
                f.write(i)
        with open('sec.txt') as f:
            tyuio = f.read()
        tyuio = tyuio.replace(self.sector,'')
        tyuio = tyuio.split(',')
        for i in tyuio:
            if i == '':
                del tyuio[tyuio.index(i)]
            else:
                tyuio[tyuio.index(i)] = tyuio[tyuio.index(i)]+','
        tyuio[-1] = tyuio[-1].strip(',')
        with open('sec.txt','w') as f:
            for i in tyuio:
                f.write(i)
